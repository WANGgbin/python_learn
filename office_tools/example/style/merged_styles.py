import openpyxl as xl
from openpyxl.styles import Border, Side, Font, Alignment
def run():
    wb = xl.Workbook()
    ws = wb.active
    ws.title = 'style_merged_cells'

    ws.merge_cells('A1:D6')
    # 合并单元格的格式通过左上角的 cell 来设置
    top_left_cell = ws['A1']
    top_left_cell.value = 'Love'
    top_left_cell.font = Font(color='FF0000')

    thin = Side(border_style='thin', color='000000')
    fat = Side(border_style='double', color='ff0000')
    top_left_cell.border = Border(left=thin, right=thin, top=fat, bottom=fat)

    # 对齐方式，两个方向。垂直方向：top|center|bottom, 水平方向：left|center|right
    top_left_cell.alignment = Alignment(vertical='top', horizontal='left')

    wb.save('merged_cells_style.xlsx')

if __name__ == '__main__':
    run()